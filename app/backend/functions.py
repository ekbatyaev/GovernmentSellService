import asyncio
import os
import pandas as pd
from datetime import datetime, timedelta
from fastapi import HTTPException, status
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert as pg_insert
from app.backend.db.table_models import AuthCode
from app.backend.db.static_info import REGIONS_OF_THE_FILTERS, ALL_REGION_CODES
from app.backend.db.table_models import Purchase
from app.backend.parsers.request_archives import get_docs_by_region, download_archive_from_result
from app.backend.parsers.xml_parser import parse_zip_archive_purchases, parse_zip_archive_protocols
from app.settings import settings, logger, MOSCOW_TZ
from typing import Dict
from app.backend.email.functions import send_email
from app.backend.api_client import api_datum_query

# Сколько регионов обрабатывать одновременно.
REGION_CONCURRENCY = 8
# Сколько архивов (суммарно по всем регионам) скачивать и парсить одновременно.
ARCHIVE_CONCURRENCY = 4
# Сколько записей (purchase/protocol, суммарно по всем архивам) одновременно
# писать через api_datum_query.
RECORD_CONCURRENCY = 10

async def set_auth_code(db: AsyncSession, email: str, code: int) -> None:
    now = datetime.now(MOSCOW_TZ).replace(tzinfo=None)
    stmt = pg_insert(AuthCode).values(email=email, code=code, created_at=now)
    stmt = stmt.on_conflict_do_update(
        index_elements=[AuthCode.email],
        set_={"code": stmt.excluded.code, "created_at": stmt.excluded.created_at},
    )
    await db.execute(stmt)
    await db.commit()

async def pop_auth_code(db: AsyncSession, email: str) -> int | None:
    """Достаёт код и сразу удаляет запись — на уровне БД это атомарно."""
    stmt = delete(AuthCode).where(AuthCode.email == email).returning(AuthCode.code, AuthCode.created_at)
    row = (await db.execute(stmt)).first()
    await db.commit()
    if row is None:
        return None
    code, created_at = row
    if datetime.now(MOSCOW_TZ).replace(tzinfo=None) - created_at > timedelta(minutes=10):
        return None
    return code

async def verify_token(token: str):
    if settings.system_token != token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

async def delete_expired(db: AsyncSession) -> int:
    today = datetime.now(MOSCOW_TZ).date()
    stmt = (
        delete(Purchase)
        .where(Purchase.submission_close_datetime.isnot(None))
        .where(func.date(Purchase.submission_close_datetime) < today)
    )
    res = await db.execute(stmt)
    return res.rowcount or 0

def _build_analysis_workbook(rows, analysis_path) -> None:
    """Синхронная сборка Excel-файла (была телом create_analysis).
    Вызывается через asyncio.to_thread, чтобы не блокировать event loop."""

    df = pd.DataFrame(rows)
    df.to_excel(analysis_path, index=False)

    wb = load_workbook(analysis_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="366092")
    header_font = Font(color="FFFFFF", bold=True)
    cell_font = Font(size=10)
    align = Alignment(wrap_text=True, vertical="center")
    thin_side = Side(style="thin")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for col in range(1, ws.max_column + 1):
        c = ws.cell(1, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align
        c.border = border

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 30
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = cell_font
            cell.alignment = align
            cell.border = border

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(
            len(str(ws.cell(r, col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    last = ws.max_row + 2
    ws.cell(last, 1, "Данные по закупкам, лотам и позициям")
    ws.merge_cells(
        start_row=last,
        start_column=1,
        end_row=last,
        end_column=ws.max_column,
    )

    footnote_cell = ws.cell(last, 1)
    footnote_cell.font = Font(italic=True, size=9, color="555555")
    footnote_cell.alignment = Alignment(horizontal="center")
    footnote_cell.border = Border(top=Side(style="thin", color="AAAAAA"))

    wb.save(analysis_path)


async def create_analysis(rows, analysis_path):
    try:
        await asyncio.to_thread(_build_analysis_workbook, rows, analysis_path)
    except Exception:
        logger.info("Произошла ошибка при создании анализа")


async def send_analysis(rows, emails, created, updated, skipped, extra_rows=None, path_name_extra="all_purchases.xlsx"):

    html_content = f"""
                    <html><body style="font-family:Arial;">
                    <h2 style="color:#2E86C1;">Уведомление о заявках с госзакупок</h2>
                    <p>Новых заявок добавлено: <b style="color:#E74C3C;font-size:18px;">{created}</b></p>
                    <p>Заявок обновлено: <b style="color:#F39C12;font-size:18px;">{updated}</b></p>
                    <p>Пропущено: <b style="color:#7F8C8D;font-size:18px;">{skipped}</b></p>
                    <hr><p style="color:#888;font-size:12px;">Это письмо сформировано автоматически, отвечать на него не нужно</p>
                    </body></html>
                    """

    attachments = []

    analysis_path = "analysis.xlsx"

    await create_analysis(rows, analysis_path=analysis_path)

    attachments.append(analysis_path)

    if extra_rows:

        await create_analysis(extra_rows, analysis_path=path_name_extra)
        attachments.append(path_name_extra)

    now = datetime.now(MOSCOW_TZ)

    subject = f"Заявки с госзакупок за {now.strftime('%d.%m.%Y')}"

    try:

        for email in emails:
            try:
                await send_email(
                    email,
                    subject,
                    html_content,
                    attachments=attachments if (created or updated) else None,
                )
            except Exception:
                logger.exception("Pipeline: не удалось отправить письмо на %s", email)

    finally:
        for path in attachments:
            if os.path.exists(path):
                os.remove(path)


def _status_from_message(message: str) -> str:
    if message == "Purchase updated":
        return "updated"
    if message == "Purchase created":
        return "created"
    return "skipped"


async def _save_purchase_record(purchase: dict, registration_numbers_dict_per_day: dict, record_semaphore: asyncio.Semaphore) -> None:
    async with record_semaphore:
        try:
            data = await api_datum_query(token=settings.system_token, endpoint="put_purchase", **purchase)
            message = data.get("message")
            status = _status_from_message(message)

            if status == "updated":
                logger.info(
                    "Pipeline: обновлена закупка через API | reg=%s | guid=%s",
                    purchase.get("registration_number"),
                    purchase.get("guid"),
                )
            elif status == "created":
                logger.info(
                    "Pipeline: создана новая закупка через API | reg=%s | guid=%s",
                    purchase.get("registration_number"),
                    purchase.get("guid"),
                )
            else:
                logger.warning(
                    "Pipeline: неизвестный ответ API | reg=%s | guid=%s | message=%s",
                    purchase.get("registration_number"),
                    purchase.get("guid"),
                    message,
                )

            registration_numbers_dict_per_day[purchase["filter_type_name"]][purchase["region_number"]][status].append(
                purchase["registration_number"]
            )

        except Exception:
            logger.exception(
                "Pipeline: ошибка сохранения закупки | reg=%s | guid=%s",
                purchase.get("registration_number"),
                purchase.get("guid"),
            )

            filter_type_name = purchase.get("filter_type_name")
            region_number = purchase.get("region_number")
            registration_number = purchase.get("registration_number")
            if filter_type_name and region_number and registration_number:
                registration_numbers_dict_per_day[filter_type_name][region_number][
                    "skipped"
                ].append(registration_number)
            else:
                logger.error(
                    "Pipeline: не удалось учесть запись как skipped — нет ключей | purchase=%s",
                    purchase,
                )


async def _save_protocol_record(protocol: dict, registration_numbers_dict_per_day: dict, record_semaphore: asyncio.Semaphore) -> None:
    async with (record_semaphore):
        try:
            database_answer = await api_datum_query(
                token=settings.system_token,
                endpoint="update_purchase",
                registration_number=protocol["registration_number"],
                result_info=protocol["result_info"],
                documents_list=protocol["documents_list"],
                publication_datetime=protocol["publication_datetime"],
            )

            if database_answer.get("message") == "Purchase not found" or database_answer == {}:
                try:
                    data = await api_datum_query(
                        token=settings.system_token, endpoint="put_purchase", **protocol
                    )
                    message = data.get("message")
                    status = _status_from_message(message)

                    if status == "updated":
                        logger.info(
                            "Pipeline: обновлена закупка через API | reg=%s | guid=%s",
                            protocol.get("registration_number"),
                            protocol.get("guid"),
                        )
                    elif status == "created":
                        logger.info(
                            "Pipeline: создана новая закупка через API | reg=%s | guid=%s",
                            protocol.get("registration_number"),
                            protocol.get("guid"),
                        )
                    else:
                        logger.warning(
                            "Pipeline: неизвестный ответ API | reg=%s | guid=%s | message=%s",
                            protocol.get("registration_number"),
                            protocol.get("guid"),
                            message,
                        )

                    registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                        status
                    ].append(protocol["registration_number"])

                    logger.info("Create purchase from protocol response.json | %s", database_answer)
                    logger.info(
                        "Pipeline: протокол создал закупку | reg=%s",
                        protocol["registration_number"],
                    )
                except Exception:
                    logger.exception(
                        "Pipeline: ошибка сохранения закупки | reg=%s | guid=%s",
                        protocol.get("registration_number"),
                        protocol.get("guid"),
                    )
                    filter_type_name = protocol.get("filter_type_name")
                    region_number = protocol.get("region_number")
                    registration_number = protocol.get("registration_number")
                    if filter_type_name and region_number and registration_number:
                        registration_numbers_dict_per_day[filter_type_name][region_number][
                            "skipped"
                        ].append(registration_number)
                    else:
                        logger.error(
                            "Pipeline: не удалось учесть запись как skipped — нет ключей | protocol=%s",
                            protocol,
                        )
                # В оригинале здесь стоял `finally: continue` — обработка
                # записи на этом заканчивалась, до кода ниже (пометка
                # "updated") очередь не доходила. return делает то же самое.
                return

            registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                "updated"
            ].append(protocol["registration_number"])

            logger.info("Update response.json | %s", database_answer)
            logger.info(
                "Pipeline: протокол обновил закупку | reg=%s",
                protocol["registration_number"],
            )

        except Exception as error:
            logger.exception(
                "У заявки не обновились поля | reg=%s | error=%s",
                protocol.get("registration_number"),
                error,
            )
            filter_type_name = protocol.get("filter_type_name")
            region_number = protocol.get("region_number")
            registration_number = protocol.get("registration_number")
            if filter_type_name and region_number and registration_number:
                registration_numbers_dict_per_day[filter_type_name][region_number][
                    "skipped"
                ].append(registration_number)
            else:
                logger.error(
                    "Pipeline: не удалось учесть запись как skipped — нет ключей | purchase=%s",
                    protocol,
                )


async def _process_purchases_for_region(
    region,
    date_str: str,
    filter_number: int,
    registration_numbers_dict_per_day: dict,
    archive_semaphore: asyncio.Semaphore,
    record_semaphore: asyncio.Semaphore,
) -> None:
    try:
        result_purchases = await get_docs_by_region(
            org_region=region,
            document_type="purchaseNotice",
            exact_date=date_str,
            subsystem_type="RI223",
        )

    except Exception:
        logger.exception("Pipeline: не удалось получить список закупок | region=%s", region)
        return

    archive_urls_purchases = result_purchases.get("archive_urls", [])

    async def _handle_archive(archive_url: str) -> None:
        async with archive_semaphore:
            try:
                zip_path_purchases = await download_archive_from_result(archive_url)

                if zip_path_purchases is None:
                    logger.info("Pipeline: скачивание архива не удалось")
                    return

                logger.info("Pipeline: архив закупок скачан: %s", zip_path_purchases)

                purchases = await parse_zip_archive_purchases(zip_path_purchases, region, filter_number)

                logger.info("Pipeline: после фильтров закупок: %s", len(purchases))

                await asyncio.gather(*[
                    _save_purchase_record(purchase, registration_numbers_dict_per_day, record_semaphore)
                    for purchase in purchases
                ])
            except Exception:
                logger.exception(
                    "Pipeline: ошибка обработки архива закупок | region=%s | archive_url=%s",
                    region,
                    archive_url,
                )

    await asyncio.gather(*[_handle_archive(url) for url in archive_urls_purchases])


async def _process_protocols_for_region(
    region,
    date_str: str,
    filter_number: int,
    registration_numbers_dict_per_day: dict,
    archive_semaphore: asyncio.Semaphore,
    record_semaphore: asyncio.Semaphore,
) -> None:
    try:
        result_protocols = await get_docs_by_region(
            org_region=region,
            document_type="purchaseProtocol",
            exact_date=date_str,
            subsystem_type="RI223",
        )
    except Exception:
        logger.exception("Pipeline: не удалось получить список протоколов | region=%s", region)
        return
    archive_urls_protocols = result_protocols.get("archive_urls", [])

    async def _handle_archive(archive_url: str) -> None:
        async with archive_semaphore:
            try:
                zip_path_protocols = await download_archive_from_result(archive_url)

                if zip_path_protocols is None:
                    logger.info("Pipeline: скачивание архива не удалось")
                    return

                logger.info("Pipeline: архив протоколов скачан: %s", zip_path_protocols)

                protocols = await parse_zip_archive_protocols(zip_path_protocols, region, filter_number)

                logger.info("Pipeline: после фильтров протоколов: %s", len(protocols))

                await asyncio.gather(*[
                    _save_protocol_record(protocol, registration_numbers_dict_per_day, record_semaphore)
                    for protocol in protocols
                ])
            except Exception:
                logger.exception(
                    "Pipeline: ошибка обработки архива протоколов | region=%s | archive_url=%s",
                    region,
                    archive_url,
                )

    await asyncio.gather(*[_handle_archive(url) for url in archive_urls_protocols])


async def process_day(date_str: str, filter_number=0) -> Dict:
    logger.info("Pipeline: обработка даты %s", date_str)

    registration_numbers_dict_per_day = {}

    for filter_name in REGIONS_OF_THE_FILTERS.keys():
        registration_numbers_dict_per_day[filter_name] = {}
        for region_code in REGIONS_OF_THE_FILTERS[filter_name]:
            registration_numbers_dict_per_day[filter_name][region_code] = {
                "created": [],
                "updated": [],
                "skipped": []
            }

    region_semaphore = asyncio.Semaphore(REGION_CONCURRENCY)
    archive_semaphore = asyncio.Semaphore(ARCHIVE_CONCURRENCY)
    record_semaphore = asyncio.Semaphore(RECORD_CONCURRENCY)

    async def _process_region(region) -> None:
        async with region_semaphore:
            try:
                # Закупки и протоколы одного региона не пересекаются по
                # данным — тянем параллельно.
                await asyncio.gather(
                    _process_purchases_for_region(
                        region, date_str, filter_number, registration_numbers_dict_per_day,
                        archive_semaphore, record_semaphore,
                    ),
                    _process_protocols_for_region(
                        region, date_str, filter_number, registration_numbers_dict_per_day,
                        archive_semaphore, record_semaphore,
                    ),
                )
            except Exception:
                logger.exception("Pipeline: не удалось обработать регион %s", region)

    await asyncio.gather(*[_process_region(region) for region in ALL_REGION_CODES])

    return registration_numbers_dict_per_day