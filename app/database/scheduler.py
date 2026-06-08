import logging
from datetime import  timezone
from time import sleep
from dateutil.relativedelta import relativedelta
from sqlalchemy import delete, func, select
from .connection_to_database import db_session
from .table_models import Purchase
from goszakupki_requests.xml_archives_request import get_docs_by_region, download_archive_from_result
from goszakupki_requests.parse_xml_archive_223fz import parse_zip_archive_purchases, parse_zip_archive_protocols
from .email_handles import send_email
from dotenv import load_dotenv
import os
import time as time_module
from datetime import datetime, timedelta, time
from typing import Dict, Any
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


load_dotenv()

logger = logging.getLogger(__name__)

LOCK_KEY = int(os.getenv("SCHEDULER_LOCK_KEY", "424242"))

RETRY_COUNT = int(os.getenv("RETRY_COUNT"))
RETRY_DELAY =  int(os.getenv("RETRY_DELAY"))
BACKFILL_DAYS = int(os.getenv("BACKFILL_DAYS", "10"))
BACKFILL_ON_STARTUP = os.getenv("PIPELINE_BACKFILL_ON_STARTUP", "true").lower() == "true"

APP_URL = os.getenv("APP_URL")
API_BASE = os.getenv("API_BASE")
TOKEN = os.getenv("SYSTEM_TOKEN")

ALL_REGION_CODES = [
    "01", "02", "03", "04", "05", "06", "07", "08", "09",
    "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
    "20", "21", "22", "23", "24", "25", "26", "27", "28", "29",
    "30", "31", "32", "33", "34", "35", "36", "37", "38", "39",
    "40", "41", "42", "43", "44", "45", "46", "47", "48", "49",
    "50", "51", "52", "53", "54", "55", "56", "57", "58", "59",
    "60", "61", "62", "63", "64", "65", "66", "67", "68", "69",
    "70", "71", "72", "73", "74", "75", "76", "77", "78", "79",
    "80", "81", "82", "83", "85", "86", "87", "89", "92", "95"
]

REGION_CODES_BY_FEDERAL_DISTRICT = {
    "Центральный федеральный округ": [
        "31",  # Белгородская область
        "32",  # Брянская область
        "33",  # Владимирская область
        "36",  # Воронежская область
        "37",  # Ивановская область
        "40",  # Калужская область
        "44",  # Костромская область
        "46",  # Курская область
        "48",  # Липецкая область
        "50",  # Московская область
        "57",  # Орловская область
        "62",  # Рязанская область
        "67",  # Смоленская область
        "68",  # Тамбовская область
        "69",  # Тверская область
        "71",  # Тульская область
        "76",  # Ярославская область
        "77",  # Москва
    ],

    "Северо-Западный федеральный округ": [
        "10",  # Республика Карелия
        "11",  # Республика Коми
        "29",  # Архангельская область
        "35",  # Вологодская область
        "39",  # Калининградская область
        "47",  # Ленинградская область
        "51",  # Мурманская область
        "53",  # Новгородская область
        "60",  # Псковская область
        "78",  # Санкт-Петербург
        "83",  # Ненецкий автономный округ
    ],

    "Южный федеральный округ": [
        "01",  # Республика Адыгея
        "08",  # Республика Калмыкия
        "23",  # Краснодарский край
        "30",  # Астраханская область
        "34",  # Волгоградская область
        "61",  # Ростовская область
        "82",  # Республика Крым
        "92",  # Севастополь
    ],

    "Северо-Кавказский федеральный округ": [
        "05",  # Республика Дагестан
        "06",  # Республика Ингушетия
        "07",  # Кабардино-Балкарская Республика
        "09",  # Карачаево-Черкесская Республика
        "15",  # Республика Северная Осетия — Алания
        "20",  # Чеченская Республика, старый код
        "26",  # Ставропольский край
        "95",  # Чеченская Республика
    ],

    "Приволжский федеральный округ": [
        "02",  # Республика Башкортостан
        "12",  # Республика Марий Эл
        "13",  # Республика Мордовия
        "16",  # Республика Татарстан
        "18",  # Удмуртская Республика
        "21",  # Чувашская Республика
        "43",  # Кировская область
        "52",  # Нижегородская область
        "56",  # Оренбургская область
        "58",  # Пензенская область
        "59",  # Пермский край
        "63",  # Самарская область
        "64",  # Саратовская область
        "73",  # Ульяновская область
        "81",  # бывший Коми-Пермяцкий АО, сейчас в составе Пермского края
    ],

    "Уральский федеральный округ": [
        "45",  # Курганская область
        "66",  # Свердловская область
        "72",  # Тюменская область
        "74",  # Челябинская область
        "86",  # Ханты-Мансийский автономный округ — Югра
        "89",  # Ямало-Ненецкий автономный округ
    ],

    "Сибирский федеральный округ": [
        "04",  # Республика Алтай
        "17",  # Республика Тыва
        "19",  # Республика Хакасия
        "22",  # Алтайский край
        "24",  # Красноярский край
        "38",  # Иркутская область
        "42",  # Кемеровская область — Кузбасс
        "54",  # Новосибирская область
        "55",  # Омская область
        "70",  # Томская область
        "85",  # бывший Усть-Ордынский Бурятский АО, сейчас в составе Иркутской области
    ],

    "Дальневосточный федеральный округ": [
        "03",  # Республика Бурятия
        "14",  # Республика Саха (Якутия)
        "25",  # Приморский край
        "27",  # Хабаровский край
        "28",  # Амурская область
        "41",  # Камчатский край
        "49",  # Магаданская область
        "65",  # Сахалинская область
        "75",  # Забайкальский край
        "79",  # Еврейская автономная область
        "80",  # бывший Агинский Бурятский АО, сейчас в составе Забайкальского края
        "87",  # Чукотский автономный округ
    ],
}


REGIONS_OF_THE_FILTERS = \
    {
        "Тендеры Россетей": ["77"],
        "Тендеры для OEM": [
    "01", "02", "03", "04", "05", "06", "07", "08", "09",
    "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
    "20", "21", "22", "23", "24", "25", "26", "27", "28", "29",
    "30", "31", "32", "33", "34", "35", "36", "37", "38", "39",
    "40", "41", "42", "43", "44", "45", "46", "47", "48", "49",
    "50", "51", "52", "53", "54", "55", "56", "57", "58", "59",
    "60", "61", "62", "63", "64", "65", "66", "67", "68", "69",
    "70", "71", "72", "73", "74", "75", "76", "77", "78", "79",
    "80", "81", "82", "83", "85", "86", "87", "89", "92", "95"
    ],
        "Тендеры для ITM": [
    "01", "02", "03", "04", "05", "06", "07", "08", "09",
    "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
    "20", "21", "22", "23", "24", "25", "26", "27", "28", "29",
    "30", "31", "32", "33", "34", "35", "36", "37", "38", "39",
    "40", "41", "42", "43", "44", "45", "46", "47", "48", "49",
    "50", "51", "52", "53", "54", "55", "56", "57", "58", "59",
    "60", "61", "62", "63", "64", "65", "66", "67", "68", "69",
    "70", "71", "72", "73", "74", "75", "76", "77", "78", "79",
    "80", "81", "82", "83", "85", "86", "87", "89", "92", "95"
    ]
    }

if not TOKEN:
    raise RuntimeError("SYSTEM_TOKEN is required")

# in-memory статус (простая телеметрия)
LAST_JOB_STATUS: Dict[str, Any] = {
    "running": False,
    "job": None,
    "started_at": None,
    "finished_at": None,
    "message": "idle",
    "progress": None,
    "result": None,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(threadName)s | %(name)s | %(message)s"
)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _set_status(**kwargs):
    LAST_JOB_STATUS.update(kwargs)


def _parse_dt(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            return None
    return None

def put_purchase_to_db(purchase: dict) -> str:
    purchase_payload = dict(purchase)
    purchase_payload["token"] = TOKEN
    registration_number = purchase_payload.get("registration_number")
    guid = purchase_payload.get("guid")

    try:
        response = requests.post(
            f"{APP_URL}{API_BASE}/put_purchase",
            json=purchase_payload,
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
    except requests.RequestException as e:
        logger.exception(
            "Pipeline: ошибка запроса к API | reg=%s | guid=%s | error=%s",
            registration_number,
            guid,
            e,
        )
        return "skipped"
    except ValueError:
        logger.warning(
            "Pipeline: API вернул некорректный JSON | reg=%s | guid=%s",
            registration_number,
            guid,
        )
        return "skipped"

    message = data.get("message")

    if message == "Purchase updated":
        logger.info(
            "Pipeline: обновлена закупка через API | reg=%s | guid=%s",
            registration_number,
            guid,
        )
        return "updated"

    if message == "Purchase created":
        logger.info(
            "Pipeline: создана новая закупка через API | reg=%s | guid=%s",
            registration_number,
            guid,
        )
        return "created"

    logger.warning(
        "Pipeline: неизвестный ответ API | reg=%s | guid=%s | message=%s",
        registration_number,
        guid,
        message,
    )
    return "skipped"

def get_last_status() -> Dict[str, Any]:
    return LAST_JOB_STATUS


def delete_expired(db) -> int:
    today = _utcnow().date()

    stmt = (
        delete(Purchase)
        .where(Purchase.submission_close_datetime.isnot(None))
        .where(func.date(Purchase.submission_close_datetime) < today)
    )

    res = db.execute(stmt)
    return res.rowcount or 0


def process_day(date_str: str, filter_number = 0) -> Dict:
    logger.info("Pipeline: обработка даты %s", date_str)

    registration_numbers_dict_per_day = {}

    for filter_name in REGIONS_OF_THE_FILTERS.keys():
        registration_numbers_dict_per_day[filter_name] = {}
        for region_code in REGIONS_OF_THE_FILTERS[filter_name]:
            registration_numbers_dict_per_day[filter_name][region_code] = {
                "created": [],
                "updated": [],
                "skipped": []
            }

    for region in ALL_REGION_CODES:

        result_purchases = get_docs_by_region(
            org_region=region,
            document_type="purchaseNotice",
            exact_date=date_str,
            subsystem_type="RI223",
        )

        archive_urls_purchases = result_purchases.get("archive_urls", [])
        for archive_url in archive_urls_purchases:
            zip_path_purchases = download_archive_from_result(archive_url)

            logger.info("Pipeline: архив закупок скачан: %s", zip_path_purchases)

            purchases = parse_zip_archive_purchases(zip_path_purchases, region, filter_number)

            logger.info("Pipeline: после фильтров закупок: %s", len(purchases))

            for purchase in purchases:
                try:
                    status = put_purchase_to_db(purchase)

                    registration_numbers_dict_per_day[purchase["filter_type_name"]][purchase["region_number"]][status].append(purchase["registration_number"])

                except Exception:
                    logger.exception(
                        "Pipeline: ошибка сохранения закупки | reg=%s | guid=%s",
                        purchase.get("registration_number"),
                        purchase.get("guid"),
                    )
                    registration_numbers_dict_per_day[purchase["filter_type_name"]][purchase["region_number"]][
                        "skipped"].append(purchase["registration_number"])


        result_protocols = get_docs_by_region(
            org_region=region,
            document_type="purchaseProtocol",
            exact_date=date_str,
            subsystem_type="RI223",
        )

        archive_urls_protocols = result_protocols.get("archive_urls", [])

        for archive_url in archive_urls_protocols:
            zip_path_protocols = download_archive_from_result(archive_url)

            logger.info("Pipeline: архив протоколов скачан: %s", zip_path_protocols)

            protocols = parse_zip_archive_protocols(zip_path_protocols, region, filter_number)

            logger.info("Pipeline: после фильтров протоколов: %s", len(protocols))

            for protocol in protocols:


                try:
                    response = requests.post(
                        f"{APP_URL}{API_BASE}/update_purchase",
                        json={
                            "token": TOKEN,
                            "registration_number": protocol["registration_number"],
                            "result_info": protocol["result_info"],
                            "documents_list": protocol["documents_list"],
                            "publication_datetime": protocol["publication_datetime"],
                        },
                        timeout=30,
                    )
                    response.raise_for_status()

                    database_answer = response.json()

                    if database_answer.get("message") == "Purchase not found" and not database_answer.get("data"):
                        try:
                            status = put_purchase_to_db(protocol)

                            registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                                status].append(protocol["registration_number"])

                            logger.info("Create purchase from protocol response | %s", response.text)

                            logger.info(
                                "Pipeline: протокол создал закупку | reg=%s",
                                protocol["registration_number"],
                            )
                        except Exception:
                            logger.exception(
                                "Pipeline: ошибка сохранения закупки | reg=%s | guid=%s",
                                protocol.get("registration_number"),
                                protocol.get("guid"),
                            )
                            registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                                "skipped"].append(protocol["registration_number"])
                        finally:
                            continue

                    registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                        "updated"].append(protocol["registration_number"])

                    logger.info("Update response | %s", response.text)

                    logger.info(
                        "Pipeline: протокол обновил закупку | reg=%s",
                        protocol["registration_number"],
                    )

                except Exception as error:
                    logger.exception(
                        "У заявки не обновились поля | reg=%s | error=%s",
                        protocol.get("registration_number"),
                        error,
                    )
                    registration_numbers_dict_per_day[protocol["filter_type_name"]][protocol["region_number"]][
                        "skipped"].append(protocol["registration_number"])

    return registration_numbers_dict_per_day

def create_analysis(rows, emails, created, updated, skipped):

    html_content = f"""
                    <html><body style="font-family:Arial;">
                    <h2 style="color:#2E86C1;">Уведомление о заявках с госзакупок</h2>
                    <p>Новых заявок добавлено: <b style="color:#E74C3C;font-size:18px;">{created}</b></p>
                    <p>Заявок обновлено: <b style="color:#F39C12;font-size:18px;">{updated}</b></p>
                    <p>Пропущено: <b style="color:#7F8C8D;font-size:18px;">{skipped}</b></p>
                    <hr><p style="color:#888;font-size:12px;">Это письмо сформировано автоматически, отвечать на него не нужно</p>
                    </body></html>
                    """

    analysis_path = "analysis.xlsx"

    try:
        df = pd.DataFrame(rows)
        df.to_excel(analysis_path, index=False)

        wb = load_workbook(analysis_path)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="366092")
        header_font = Font(color="FFFFFF", bold=True)
        cell_font = Font(size=10)
        align = Alignment(wrap_text=True, vertical="center")
        thin_side = Side(style="thin")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        for col in range(1, ws.max_column + 1):
            c = ws.cell(1, col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = align
            c.border = border

        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 30
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font = cell_font
                cell.alignment = align
                cell.border = border

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(
                len(str(ws.cell(r, col).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

        last = ws.max_row + 2
        ws.cell(last, 1, "Сноска: данные по закупкам, лотам и позициям")
        ws.merge_cells(
            start_row=last,
            start_column=1,
            end_row=last,
            end_column=ws.max_column,
        )

        footnote_cell = ws.cell(last, 1)
        footnote_cell.font = Font(italic=True, size=9, color="555555")
        footnote_cell.alignment = Alignment(horizontal="center")
        footnote_cell.border = Border(top=Side(style="thin", color="AAAAAA"))

        wb.save(analysis_path)

        now = datetime.now()

        subject = f"Заявки с госзакупок за {now.strftime('%d.%m.%Y')}"

        for u in emails:
            email = u.get("email")
            if not email:
                continue

            send_email(
                email,
                subject,
                html_content,
                attachments=[analysis_path] if (created or updated) else None,
            )

    finally:
        if os.path.exists(analysis_path):
            os.remove(analysis_path)

def run_daily_job() -> Dict[str, Any]:
    _set_status(
        running=True,
        job="daily",
        started_at=_utcnow().isoformat(),
        finished_at=None,
        message="running",
        progress=None,
        result=None,
    )

    try:
        yesterday = (_utcnow() - timedelta(days=1)).date()
        date_str = yesterday.strftime("%Y-%m-%d")

        _set_status(progress={"date": date_str})

        max_attempts = 3
        retry_delay_sec = 10
        day_result = None

        for attempt in range(1, max_attempts + 1):
            try:
                logger.info(
                    "Scheduler: daily attempt %s/%s for date %s",
                    attempt,
                    max_attempts,
                    date_str,
                )
                day_result = process_day(date_str, 0)
                break
            except Exception as e:
                logger.warning(
                    "Scheduler: daily error (%s) attempt %s/%s: %s",
                    date_str,
                    attempt,
                    max_attempts,
                    e,
                )
                if attempt == max_attempts:
                    raise
                time_module.sleep(retry_delay_sec)

        if day_result is None:
            raise RuntimeError("Daily job finished without result")


        now = datetime.now()

        # Получение старых заявок и их удаление

        delete_date_to = now - relativedelta(years=1)

        purchases_response = requests.post(
            f"{APP_URL}{API_BASE}/get_all_purchases",
            json={"token": TOKEN,
                  "publication_datetime_to": delete_date_to.replace(microsecond=0).isoformat()},
            timeout=30,
        )

        purchases_response.raise_for_status()
        purchases = purchases_response.json().get("data", [])

        count = 0
        for purchase in purchases:
            try:
                delete_response = requests.post(
                    f"{APP_URL}{API_BASE}/delete_purchase",
                    json={"token": TOKEN, "guid": purchase["guid"]},
                    timeout=30,
                )

                count += 1

                delete_response.raise_for_status()
            except Exception as e:
                logger.info(
                    f"При удалении {purchase["guid"]}, произошла ошибка {e}",
                )
                continue

        logger.info(
            f"Daily job: успешно удалено {count} заявок"
        )

        # Рассылка сообщений адрессантам

        # Высылаем информацию по Россетям

        result = {
            "ok": True,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "date": date_str,
        }

        created = 0
        updated = 0
        skipped = 0

        registration_numbers = []

        filter_name = "Тендеры Россетей"

        for region in day_result[filter_name].keys():
            created += len(day_result[filter_name][region]["created"])
            updated += len(day_result[filter_name][region]["updated"])
            skipped += len(day_result[filter_name][region]["skipped"])

            registration_numbers += (day_result[filter_name][region]["created"] +
                                     day_result[filter_name][region]["updated"] +
                                     day_result[filter_name][region]["skipped"])

        emails_response = requests.post(
            f"{APP_URL}{API_BASE}/get_all_newsletters",
            json={"token": TOKEN, "filter_type_name": filter_name},
            timeout=30,
        )

        emails_response.raise_for_status()
        emails = emails_response.json().get("data", [])

        rows = []

        for registration_number in registration_numbers:
            purchase_response = requests.post(
                f"{APP_URL}{API_BASE}/get_purchase",
                json={"token": TOKEN, "registration_number": registration_number},
                timeout=30,
            )

            purchase_response.raise_for_status()

            purchase = purchase_response.json().get("data", {})

            customer = purchase.get("customer") or {}
            result_info = purchase.get("result_info") or {}

            base = {
                "Реестровый номер": purchase.get("registration_number"),
                "Название закупки": purchase.get("name"),
                "Сумма закупки": purchase.get("initial_sum"),
                "Дата начала подачи заявок": purchase.get("submission_start_datetime"),
                "Дата окончания подачи заявок": purchase.get("submission_close_datetime"),
                "Дата публикации": purchase.get("publication_datetime"),
                "Заказчик название": customer.get("full_name"),
                "Победитель": result_info.get("Победитель"),
                "Другие участники": result_info.get("Другие участники"),
                "Ячейки": result_info.get("Ячейки"),
                "Кол-во ячеек": result_info.get("Кол-во ячеек"),
                "Типовой проект": result_info.get("Типовой проект"),
                "Проектировщик": result_info.get("Проектировщик"),
                "Дата исполнения договора": result_info.get("Дата исполнения договора"),
                "Филиал/РЭС": result_info.get("Филиал/РЭС"),
                "Ссылка на тендер": f"https://zakupki.gov.ru/epz/order/notice/notice223/common-info.html?regNumber={purchase.get("registration_number")}"
            }

            rows.append(base)

        result["created"] += created
        result["updated"] += updated
        result["skipped"] += skipped


        create_analysis(rows=rows, emails=emails,
                        created=created, updated=updated, skipped=skipped)


        # Высылаем информацию для OEM

        filter_name = "Тендеры для OEM"

        for district in REGION_CODES_BY_FEDERAL_DISTRICT.keys():

            created = 0
            updated = 0
            skipped = 0

            registration_numbers = []

            for region in REGION_CODES_BY_FEDERAL_DISTRICT[district]:
                created += len(day_result[filter_name][region]["created"])
                updated += len(day_result[filter_name][region]["updated"])
                skipped += len(day_result[filter_name][region]["skipped"])

                registration_numbers += (day_result[filter_name][region]["created"] +
                                         day_result[filter_name][region]["updated"] +
                                         day_result[filter_name][region]["skipped"])

            emails_response = requests.post(
                f"{APP_URL}{API_BASE}/get_all_newsletters",
                json={"token": TOKEN, "filter_type_name": filter_name, "district_name": district},
                timeout=30,
            )

            emails_response.raise_for_status()
            emails = emails_response.json().get("data", [])

            rows = []

            for registration_number in registration_numbers:
                purchase_response = requests.post(
                    f"{APP_URL}{API_BASE}/get_purchase",
                    json={"token": TOKEN, "registration_number": registration_number},
                    timeout=30,
                )

                purchase_response.raise_for_status()

                purchase = purchase_response.json().get("data", {})

                customer = purchase.get("customer") or {}

                base = {
                    "Реестровый номер": purchase.get("registration_number"),
                    "Название закупки": purchase.get("name"),
                    "Сумма закупки": purchase.get("initial_sum"),
                    "Дата начала подачи заявок": purchase.get("submission_start_datetime"),
                    "Дата окончания подачи заявок": purchase.get("submission_close_datetime"),
                    "Дата публикации": purchase.get("publication_datetime"),
                    "Заказчик название": customer.get("full_name"),
                    "Ссылка на тендер": f"https://zakupki.gov.ru/epz/order/notice/notice223/common-info.html?regNumber={purchase.get("registration_number")}"
                }

                rows.append(base)

            result["created"] += created
            result["updated"] += updated
            result["skipped"] += skipped

            create_analysis(rows=rows, emails=emails,
                            created=created, updated=updated, skipped=skipped)

        # Высылаем информацию для ITM

        filter_name = "Тендеры для ITM"

        for district in REGION_CODES_BY_FEDERAL_DISTRICT.keys():

            created = 0
            updated = 0
            skipped = 0

            registration_numbers = []

            for region in REGION_CODES_BY_FEDERAL_DISTRICT[district]:
                created += len(day_result[filter_name][region]["created"])
                updated += len(day_result[filter_name][region]["updated"])
                skipped += len(day_result[filter_name][region]["skipped"])

                registration_numbers += (day_result[filter_name][region]["created"] +
                                         day_result[filter_name][region]["updated"] +
                                         day_result[filter_name][region]["skipped"])

            emails_response = requests.post(
                f"{APP_URL}{API_BASE}/get_all_newsletters",
                json={"token": TOKEN, "filter_type_name": filter_name},
                timeout=30,
            )

            emails_response.raise_for_status()
            emails = emails_response.json().get("data", [])

            rows = []

            for registration_number in registration_numbers:
                purchase_response = requests.post(
                    f"{APP_URL}{API_BASE}/get_purchase",
                    json={"token": TOKEN, "registration_number": registration_number},
                    timeout=30,
                )

                purchase_response.raise_for_status()

                purchase = purchase_response.json().get("data", {})

                customer = purchase.get("customer") or {}

                result_info = purchase.get("result_info") or {}

                base = {
                    "Реестровый номер": purchase.get("registration_number"),
                    "Название закупки": purchase.get("name"),
                    "Сумма закупки": purchase.get("initial_sum"),
                    "Дата начала подачи заявок": purchase.get("submission_start_datetime"),
                    "Дата окончания подачи заявок": purchase.get("submission_close_datetime"),
                    "Дата публикации": purchase.get("publication_datetime"),
                    "Заказчик название": customer.get("full_name"),
                    "Победитель": result_info.get("Победитель"),
                    "ИНН": result_info.get("ИНН"),
                    "Итоговая цена контракта": result_info.get("Итоговая цена контракта"),
                    "Другие участники": result_info.get("Другие участники"),
                    "Ссылка на тендер": f"https://zakupki.gov.ru/epz/order/notice/notice223/common-info.html?regNumber={purchase.get("registration_number")}"
                }

                rows.append(base)

            result["created"] += created
            result["updated"] += updated
            result["skipped"] += skipped

            create_analysis(rows=rows, emails=emails,
                            created=created, updated=updated, skipped=skipped)


        _set_status(
            running=False,
            finished_at=_utcnow().isoformat(),
            message="done",
            result=result,
        )
        logger.info("Scheduler: daily done | %s", result)
        return result

    except Exception as e:
        logger.exception("Scheduler: daily job failed")
        error_result = {"ok": False, "error": str(e)}

        _set_status(
            running=False,
            finished_at=_utcnow().isoformat(),
            message=f"error: {e}",
            result=error_result,
        )
        return error_result


def run_backfill(days: int | None = None, filter_number = 0) -> Dict[str, Any]:
    days = int(days or BACKFILL_DAYS)
    _set_status(
        running=True,
        job="backfill",
        started_at=_utcnow().isoformat(),
        finished_at=None,
        message="running",
        progress={"days": days, "current": 0, "date": None, "created_total": 0, "updated_total": 0},
        result=None,
    )

    created_total = 0
    updated_total = 0
    skipped_total = 0
    processed_days = 0
    failed_days = []

    try:
        today = _utcnow().date()
        start = today - timedelta(days=days)

        for i in range(days):
            d = start + timedelta(days=i)
            if d >= today:
                continue

            date_str = d.strftime("%Y-%m-%d")
            processed_days += 1

            _set_status(progress={
                "days": days,
                "current": i + 1,
                "date": date_str,
                "created_total": created_total,
                "updated_total": updated_total,
                "skipped_total": skipped_total,
            })

            logger.info("Backfill: day %s/%s | %s", i + 1, days, date_str)

            success = False
            for attempt in range(RETRY_COUNT):
                try:
                    day_result = process_day(date_str, filter_number)
                    for filter_name in REGIONS_OF_THE_FILTERS:

                        for region in day_result[filter_name].keys():
                            created_total += len(day_result[filter_name][region]["created"])
                            updated_total += len(day_result[filter_name][region]["updated"])
                            skipped_total += len(day_result[filter_name][region]["skipped"])

                    success = True
                    break
                except Exception as e:
                    logger.warning(
                        "Backfill error (%s) attempt %s/%s: %s",
                        date_str, attempt + 1, RETRY_COUNT, e
                    )
                    sleep(RETRY_DELAY)

            if not success:
                logger.error("Backfill: пропускаю дату %s после %s попыток", date_str, RETRY_COUNT)
                failed_days.append(date_str)

        result = {
            "ok": True,
            "processed_days": processed_days,
            "created_total": created_total,
            "updated_total": updated_total,
            "skipped_total": skipped_total,
            "failed_days": failed_days,
        }

        _set_status(
            running=False,
            finished_at=_utcnow().isoformat(),
            message="done",
            result=result
        )

        logger.info("Scheduler(backfill): done | %s", result)
        return result

    except Exception as e:
        logger.exception("Scheduler(backfill) failed")
        _set_status(
            running=False,
            finished_at=_utcnow().isoformat(),
            message=f"error: {e}",
            result={"ok": False, "error": str(e)}
        )
        return {"ok": False, "error": str(e)}


def run_backfill_on_startup() -> None:
    if not BACKFILL_ON_STARTUP:
        logger.info("Backfill on startup disabled")
        return
    run_backfill(BACKFILL_DAYS)