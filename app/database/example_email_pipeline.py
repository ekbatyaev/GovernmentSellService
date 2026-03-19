import os, datetime, requests, pandas as pd
from dotenv import load_dotenv
from email_handles import send_email
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

APP_URL = os.getenv("APP_URL")
TOKEN = os.getenv("SYSTEM_TOKEN") or exit("SYSTEM_TOKEN is required")

added = 5
now = datetime.datetime.now()
start_day = datetime.datetime.combine(now.date(), datetime.time.min)

emails = requests.post(f"{APP_URL}/get_all_newsletters", json={"token": TOKEN}).json().get("data", [])

html_content = f"""
<html><body style="font-family:Arial;">
<h2 style="color:#2E86C1;">Уведомление о новых заявках с госзакупок</h2>
<p>Добавлено новых заявок: <b style="color:#E74C3C;font-size:18px;">{added}</b></p>
<hr><p style="color:#888;font-size:12px;">Это письмо сформировано автоматически, отвечать на него не нужно</p>
</body></html>
"""

data = requests.post(
    f"{APP_URL}/get_all_purchases",
    json={"token": TOKEN, "created_at_from": start_day.isoformat()}
).json().get("data", [])

rows = []
for p in data:
    base = {
        'guid_закупки': p['guid'],
        'reg_number': p['registration_number'],
        'название_закупки': p['name'],
        'файл_источник': p['source_file'],
        'сумма_общая': p['initial_sum'],
        'дата_публикации': p['publication_datetime'],
        'дата_окончания': p['submission_close_datetime'],
        'заказчик_инн': p['customer']['inn'],
        'заказчик_кпп': p['customer']['kpp'],
        'заказчик_огрн': p['customer']['ogrn'],
        'заказчик_название': p['customer']['full_name'],
        'контакт_email': p['contact']['email'],
        'контакт_телефон': p['contact']['phone'],
        'контакт_фио': " ".join(filter(None, [
            p['contact']['last_name'],
            p['contact']['first_name'],
            p['contact']['middle_name']
        ])),
        'порядок_подачи': p['apply_request']['submission_order'],
        'место_подачи': p['apply_request']['submission_place'],
        'дата_начала_подачи': p['apply_request']['submission_start_date'],
    }

    for lot in p['lots']:
        for item in lot.get('items', [{}]):
            rows.append({
                **base,
                'лот_guid': lot['guid'],
                'лот_номер': lot['ordinal_number'],
                'лот_предмет': lot['subject'],
                'лот_валюта': lot['currency'],
                'лот_сумма': lot['initial_sum'],
                'позиция_количество': item.get('qty'),
                'позиция_guid': item.get('guid'),
                'окпд2_код': item.get('okpd2_code'),
                'окпд2_название': item.get('okpd2_name'),
                'доп_инфо': item.get('additional_info'),
            })

df = pd.DataFrame(rows)
df.to_excel('analysis.xlsx', index=False)

wb = load_workbook('analysis.xlsx')
ws = wb.active

# стили
header_fill = PatternFill("solid", fgColor="366092")
header_font = Font(color='FFFFFF', bold=True)
cell_font = Font(size=10)
align = Alignment(wrap_text=True, vertical='center')
border = Border(*(Side(style='thin') for _ in range(4)))

for col in range(1, ws.max_column + 1):
    c = ws.cell(1, col)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, align, border

for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 30
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        cell.font, cell.alignment, cell.border = cell_font, align, border

for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    max_len = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
    ws.column_dimensions[letter].width = min(max_len + 2, 50)

# сноска
last = ws.max_row + 2
ws.cell(last, 1, 'Сноска: данные по закупкам, лотам и позициям')
ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=ws.max_column)

f = ws.cell(last, 1)
f.font = Font(italic=True, size=9, color='555555')
f.alignment = Alignment(horizontal='center')
f.border = Border(top=Side(style='thin', color='AAAAAA'))

wb.save('analysis.xlsx')

print("Файл создан")

subject = f"Заявки с госзакупок за {now.strftime('%d.%m.%Y')}"

for u in emails:
    if u.get("email"):
        send_email(u["email"], subject, html_content,
                   attachments=["analysis.xlsx"] if added else None)

os.remove('analysis.xlsx')