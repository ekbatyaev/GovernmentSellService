import os
import re
import zipfile
import shutil
import mimetypes
import tempfile
import subprocess
import hashlib
import logging
import threading
from pathlib import Path
from urllib.parse import urlparse, unquote
from typing import Dict, Optional, Iterable
import time
from contextlib import contextmanager
from collections import OrderedDict

import rarfile
import requests
import docx
from openpyxl import load_workbook
from pypdf import PdfReader
from charset_normalizer import from_bytes
from docx.table import Table
from docx.text.paragraph import Paragraph
from .extractor_pipeline import extract_tender_fields

logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = (15, 90)
CHUNK_SIZE = 1024 * 1024

MAX_TOP_LEVEL_WORKERS = 1
MAX_ARCHIVE_FILE_WORKERS = 8
MAX_PDF_PAGES = 30
MAX_TEXT_CHARS = 300_000
MAX_TEXT_FILE_BYTES = 10 * 1024 * 1024
LIBREOFFICE_MAX_PARALLEL = 1
LIBREOFFICE_CONVERT_TIMEOUT = 20

SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".docx", ".doc", ".xlsx", ".xls"}
ARCHIVE_EXTENSIONS = {".zip", ".rar"}

RAR_PART_RE = re.compile(r"^(?P<base>.+)\.part(?P<part>\d+)\.rar$", re.IGNORECASE)
RAR_OLDSTYLE_MAIN_RE = re.compile(r"^(?P<base>.+)\.rar$", re.IGNORECASE)
RAR_OLDSTYLE_SUBPART_RE = re.compile(r"^(?P<base>.+)\.r(?P<part>\d{2})$", re.IGNORECASE)

RAR_TOOL = shutil.which("unrar")
if not RAR_TOOL:
    raise RuntimeError("Не найден системный инструмент unrar")

rarfile.UNRAR_TOOL = "unrar"

_thread_local = threading.local()
_libreoffice_semaphore = threading.Semaphore(LIBREOFFICE_MAX_PARALLEL)

FIELDS = [
    "Победитель",
    "Другие участники",
    "Ячейки",
    "Кол-во ячеек",
    "Типовой проект",
    "Проектировщик",
    "Дата исполнения договора",
    "Филиал/РЭС",
]


# =========================
# Утилиты
# =========================

def format_log_kv(**kwargs) -> str:
    parts = []
    for k, v in kwargs.items():
        if v is None:
            continue
        parts.append(f"{k}={v}")
    return " | ".join(parts)


def short_id(value: str) -> str:
    return hashlib.md5(value.encode("utf-8", errors="ignore")).hexdigest()[:8]


@contextmanager
def log_timed(stage: str, **kwargs):
    started = time.perf_counter()
    logger.info("%s START | %s", stage, format_log_kv(**kwargs))
    try:
        yield
    except Exception:
        elapsed = time.perf_counter() - started
        logger.exception("%s ERROR | %s | elapsed=%.3fs", stage, format_log_kv(**kwargs), elapsed)
        raise
    else:
        elapsed = time.perf_counter() - started
        logger.info("%s DONE | %s | elapsed=%.3fs", stage, format_log_kv(**kwargs), elapsed)


def is_rar_part_filename(name: str) -> bool:
    name = (name or "").strip()
    return bool(RAR_PART_RE.match(name) or RAR_OLDSTYLE_SUBPART_RE.match(name))


def get_rar_group_info(filename: str) -> Optional[dict]:
    filename = (filename or "").strip()

    m = RAR_PART_RE.match(filename)
    if m:
        return {
            "scheme": "part",
            "base": m.group("base"),
            "part_num": int(m.group("part")),
            "is_start": int(m.group("part")) == 1,
        }

    m = RAR_OLDSTYLE_SUBPART_RE.match(filename)
    if m:
        return {
            "scheme": "oldstyle_subpart",
            "base": m.group("base"),
            "part_num": int(m.group("part")) + 1,
            "is_start": False,
        }

    m = RAR_OLDSTYLE_MAIN_RE.match(filename)
    if m:
        return {
            "scheme": "oldstyle_main",
            "base": m.group("base"),
            "part_num": 0,
            "is_start": True,
        }

    return None


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
        "Referer": "https://zakupki.gov.ru/",
        "Accept": "*/*",
        "Connection": "keep-alive",
    })

    adapter = requests.adapters.HTTPAdapter(pool_connections=32, pool_maxsize=32, max_retries=2)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def get_thread_session() -> requests.Session:
    if not hasattr(_thread_local, "session"):
        _thread_local.session = make_session()
    return _thread_local.session


def ensure_dir(path: str | Path) -> Path:
    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_filename(name: str) -> str:
    name = os.path.basename((name or "").strip().replace("\x00", ""))
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    return name or "downloaded_file"


def stable_id(value: str) -> str:
    return hashlib.md5(value.encode("utf-8", errors="ignore")).hexdigest()


def read_head(path: Path, n: int = 512) -> bytes:
    with open(path, "rb") as f:
        return f.read(n)


def guess_extension_from_magic(data: bytes, original_name: str = "") -> Optional[str]:
    name = (original_name or "").lower()

    if data.startswith(b"%PDF"):
        return ".pdf"

    if data.startswith(b"Rar!\x1A\x07\x00") or data.startswith(b"Rar!\x1A\x07\x01\x00"):
        return ".rar"

    if data.startswith(b"PK\x03\x04"):
        if name.endswith(".docx"):
            return ".docx"
        if name.endswith(".xlsx"):
            return ".xlsx"
        if name.endswith(".pptx"):
            return ".pptx"
        return ".zip"

    return None


def is_probably_html(data: bytes) -> bool:
    head = data[:4096].lower()
    return b"<html" in head or b"<!doctype html" in head or b"<body" in head


def is_binary_bytes(data: bytes) -> bool:
    if not data:
        return False
    if b"\x00" in data[:4096]:
        return True
    text_like = sum(32 <= b <= 126 or b in b"\r\n\t\f\b" or b >= 128 for b in data[:4096])
    return (text_like / max(1, len(data[:4096]))) < 0.75


def detect_encoding(data: bytes) -> str:
    best = from_bytes(data).best()
    if best and best.encoding:
        return best.encoding

    for enc in ("utf-8", "utf-8-sig", "cp1251", "utf-16", "latin-1"):
        try:
            data.decode(enc)
            return enc
        except Exception:
            pass
    return "utf-8"


def normalize_merged_value(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def split_merged_value(value: str):
    parts = []
    for chunk in str(value).replace("\n", ";").split(";"):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    return parts


def init_result_accumulator() -> dict:
    return {field: OrderedDict() for field in FIELDS}


def merge_extracted_into_accumulator(accumulator: dict, extracted: dict) -> None:
    if not extracted:
        return

    for field, value in extracted.items():
        value = normalize_merged_value(value)
        if not value:
            continue

        for part in split_merged_value(value):
            part = normalize_merged_value(part)
            if part:
                accumulator[field][part] = None


def finalize_result_accumulator(accumulator: dict) -> dict:
    return {
        field: "; ".join(values.keys()) if values else ""
        for field, values in accumulator.items()
    }


def run_libreoffice_convert(src: Path, outdir: Path, target_ext: str, task_id: str | None = None) -> Path:
    with _libreoffice_semaphore:
        env = os.environ.copy()
        env["HOME"] = str(outdir)

        profile_dir = outdir / "lo_profile"
        profile_dir.mkdir(parents=True, exist_ok=True)

        cmd = [
            "libreoffice",
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--nodefault",
            "--norestore",
            f"-env:UserInstallation=file://{profile_dir.resolve()}",
            "--convert-to", target_ext,
            "--outdir", str(outdir),
            str(src),
        ]

        with log_timed(
            "LIBREOFFICE_CONVERT",
            task_id=task_id,
            src=src.name,
            src_path=src,
            target_ext=target_ext,
            outdir=outdir,
            timeout=LIBREOFFICE_CONVERT_TIMEOUT,
        ):
            try:
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    env=env,
                    timeout=LIBREOFFICE_CONVERT_TIMEOUT,
                )
            except subprocess.TimeoutExpired as e:
                raise RuntimeError(
                    f"LibreOffice timeout after {LIBREOFFICE_CONVERT_TIMEOUT}s "
                    f"for file {src.name}"
                ) from e

            converted = outdir / f"{src.stem}.{target_ext}"
            if converted.exists():
                return converted

            stderr = (result.stderr or "").strip()
            stdout = (result.stdout or "").strip()

            raise RuntimeError(
                f"Не удалось конвертировать {src.suffix.lower()} -> {target_ext}. "
                f"returncode={result.returncode}; stdout={stdout}; stderr={stderr}"
            )

# =========================
# Скачивание
# =========================

def guess_filename(response: requests.Response, url: str, fallback: str = "downloaded_file") -> str:
    cd = response.headers.get("Content-Disposition", "")
    match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, flags=re.IGNORECASE)
    if match:
        return safe_filename(unquote(match.group(1)))

    parsed = urlparse(response.url or url)
    name = Path(unquote(parsed.path)).name
    if name:
        return safe_filename(name)

    content_type = (response.headers.get("Content-Type") or "").split(";")[0].strip().lower()
    ext = mimetypes.guess_extension(content_type) or ""
    return safe_filename(fallback + ext)


def download_file(
    session: requests.Session,
    url: str,
    download_dir: str | Path,
    filename: Optional[str] = None,
    task_id: str | None = None,
) -> Path:
    download_dir = ensure_dir(download_dir)

    started = time.perf_counter()
    logger.info(
        "DOWNLOAD START | %s",
        format_log_kv(task_id=task_id, url=url, filename=filename, download_dir=download_dir)
    )

    with session.get(url, stream=True, allow_redirects=True, timeout=REQUEST_TIMEOUT) as response:
        response.raise_for_status()

        raw_name = filename or guess_filename(response, url)
        path = download_dir / safe_filename(raw_name)

        expected_size = response.headers.get("Content-Length")
        written = 0

        with open(path, "wb") as f:
            for chunk in response.iter_content(chunk_size=CHUNK_SIZE):
                if chunk:
                    f.write(chunk)
                    written += len(chunk)

    if expected_size:
        try:
            if written != int(expected_size):
                raise RuntimeError(f"Файл скачан не полностью: expected={expected_size}, got={written}")
        except ValueError:
            pass

    if not path.exists() or path.stat().st_size == 0:
        raise RuntimeError("Файл не скачан или пустой")

    data_head = read_head(path, 512)

    if is_probably_html(data_head):
        raise RuntimeError("Сервер вернул HTML вместо документа")

    magic_ext = guess_extension_from_magic(data_head[:16], path.name)
    if magic_ext and path.suffix.lower() != magic_ext and not is_rar_part_filename(path.name):
        fixed_path = path.with_suffix(magic_ext)
        path.rename(fixed_path)
        path = fixed_path

    elapsed = time.perf_counter() - started
    logger.info(
        "DOWNLOAD DONE | %s",
        format_log_kv(
            task_id=task_id,
            url=url,
            saved_as=path.name,
            saved_path=path,
            size_bytes=written,
            expected_size=expected_size,
            elapsed=f"{elapsed:.3f}s",
        )
    )

    return path


# =========================
# Чтение файлов
# =========================

def read_text_file(path: Path) -> str:
    if path.stat().st_size > MAX_TEXT_FILE_BYTES:
        raise ValueError(f"TXT слишком большой: {path.stat().st_size} байт")

    with open(path, "rb") as f:
        data = f.read()

    if is_binary_bytes(data[:4096]):
        raise ValueError("Файл бинарный, это не txt")

    enc = detect_encoding(data)
    return data.decode(enc, errors="replace").strip()


def read_pdf_file(path: Path) -> str:
    logger.debug("PDF READ TRY | %s", format_log_kv(path=path, reader="pymupdf"))
    text = read_pdf_file_pymupdf(path)
    if text.strip():
        logger.debug("PDF READ SUCCESS | %s", format_log_kv(path=path, reader="pymupdf", text_len=len(text)))
        return text

    logger.debug("PDF READ FALLBACK | %s", format_log_kv(path=path, from_reader="pymupdf", to_reader="pypdf"))
    text = read_pdf_file_pypdf(path)
    if text.strip():
        logger.debug("PDF READ SUCCESS | %s", format_log_kv(path=path, reader="pypdf", text_len=len(text)))
        return text

    raise ValueError("Текст не извлечен")


def read_pdf_file_pypdf(path: Path) -> str:
    reader = PdfReader(str(path))
    parts = []
    total_chars = 0

    for i, page in enumerate(reader.pages):
        if i >= MAX_PDF_PAGES or total_chars >= MAX_TEXT_CHARS:
            break

        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""

        if text.strip():
            parts.append(text)
            total_chars += len(text)

    return "\n".join(parts).strip()


def read_pdf_file_pymupdf(path: Path) -> str:
    try:
        import fitz
    except Exception:
        logger.debug("PDF READ SKIP | %s", format_log_kv(path=path, reason="fitz_not_installed"))
        return ""

    parts = []
    total_chars = 0
    doc = fitz.open(str(path))
    try:
        for i, page in enumerate(doc):
            if i >= MAX_PDF_PAGES or total_chars >= MAX_TEXT_CHARS:
                break

            try:
                text = page.get_text("text") or ""
            except Exception:
                text = ""

            if text.strip():
                parts.append(text)
                total_chars += len(text)
    finally:
        doc.close()

    return "\n".join(parts).strip()


def iter_block_items(parent):
    from docx.oxml.text.paragraph import CT_P
    from docx.oxml.table import CT_Tbl

    parent_elm = parent.element.body

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def read_docx_file(path: Path) -> str:
    d = docx.Document(str(path))
    parts = []

    for block in iter_block_items(d):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if text:
                parts.append(text)

        elif isinstance(block, Table):
            for row in block.rows:
                cells = []
                for cell in row.cells:
                    cell_text = "\n".join(
                        p.text.strip()
                        for p in cell.paragraphs
                        if p.text and p.text.strip()
                    ).strip()
                    cells.append(cell_text)

                row_text = "\t".join(c for c in cells if c)
                if row_text.strip():
                    parts.append(row_text)

    return "\n".join(parts).strip()


def read_doc_file(path: Path, task_id: str | None = None) -> str:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        converted = run_libreoffice_convert(path, tmp_dir, "docx", task_id=task_id)
        return read_docx_file(converted)


def read_xlsx_file(path: Path) -> str:
    parts = []
    wb = load_workbook(filename=path, read_only=True, data_only=True)

    try:
        for ws in wb.worksheets:
            parts.append(f"[Лист: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(
                    str(cell).strip()
                    for cell in row
                    if cell is not None and str(cell).strip()
                )
                if row_text:
                    parts.append(row_text)
    finally:
        wb.close()

    return "\n".join(parts).strip()


def read_xls_file(path: Path, task_id: str | None = None) -> str:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        converted = run_libreoffice_convert(path, tmp_dir, "xlsx", task_id=task_id)
        return read_xlsx_file(converted)


def read_file(path: Path, task_id: str | None = None) -> str:
    ext = path.suffix.lower()
    logger.debug("READ_DISPATCH | %s", format_log_kv(task_id=task_id, path=path, ext=ext))

    if ext == ".txt":
        return read_text_file(path)
    if ext == ".pdf":
        return read_pdf_file(path)
    if ext == ".docx":
        return read_docx_file(path)
    if ext == ".doc":
        return read_doc_file(path, task_id=task_id)
    if ext == ".xlsx":
        return read_xlsx_file(path)
    if ext == ".xls":
        return read_xls_file(path, task_id=task_id)

    raise ValueError(f"Пропущен неподдерживаемый формат: {ext or path.name}")


# =========================
# Архивы
# =========================

def extract_zip(path: Path, extract_to: Path, task_id: str | None = None) -> Path:
    started = time.perf_counter()
    logger.info("EXTRACT_ZIP START | %s", format_log_kv(task_id=task_id, path=path, extract_to=extract_to))

    with zipfile.ZipFile(path, "r") as zf:
        members = zf.namelist()
        logger.info("EXTRACT_ZIP CONTENTS | %s", format_log_kv(task_id=task_id, path=path.name, files=len(members)))
        zf.extractall(extract_to)

    elapsed = time.perf_counter() - started
    logger.info("EXTRACT_ZIP DONE | %s", format_log_kv(task_id=task_id, path=path.name, files=len(members), elapsed=f"{elapsed:.3f}s"))
    return extract_to


def extract_rar(path: Path, extract_to: Path, task_id: str | None = None) -> Path:
    started = time.perf_counter()
    logger.info("EXTRACT_RAR START | %s", format_log_kv(task_id=task_id, path=path, extract_to=extract_to))

    if not shutil.which("unrar"):
        raise RuntimeError("Не найден системный инструмент для распаковки RAR (unrar)")

    try:
        with rarfile.RarFile(path) as rf:
            if rf.needs_password():
                raise RuntimeError("RAR-архив защищён паролем")

            members = rf.infolist()
            logger.info("EXTRACT_RAR CONTENTS | %s", format_log_kv(task_id=task_id, path=path.name, files=len(members)))
            rf.extractall(path=extract_to)

    except rarfile.BadRarFile as e:
        raise RuntimeError(f"RAR-архив повреждён или имеет неподдерживаемый формат: {e}")
    except rarfile.NeedFirstVolume:
        raise RuntimeError("Для распаковки многотомного RAR нужен первый том архива")
    except rarfile.PasswordRequired:
        raise RuntimeError("RAR-архив защищён паролем")
    except rarfile.RarCannotExec as e:
        raise RuntimeError(f"Не удалось запустить unrar: {e}")
    except Exception as e:
        raise RuntimeError(f"Не удалось распаковать RAR: {e}")

    elapsed = time.perf_counter() - started
    logger.info("EXTRACT_RAR DONE | %s", format_log_kv(task_id=task_id, path=path.name, files=len(members), elapsed=f"{elapsed:.3f}s"))
    return extract_to


# =========================
# Основной reader
# =========================

def is_office_zip(path: Path) -> bool:
    ext = path.suffix.lower()
    return ext in {".docx", ".xlsx", ".pptx"}


def iter_candidate_files(target: Path) -> Iterable[Path]:
    for p in sorted(target.rglob("*")):
        if not p.is_file():
            continue

        ext = p.suffix.lower()
        if ext in SUPPORTED_EXTENSIONS or ext in ARCHIVE_EXTENSIONS:
            yield p


def process_text_into_accumulator(
    text: str,
    accumulator: dict,
    filename: str,
    task_id: str | None = None,
) -> bool:
    if not text or not text.strip():
        return False

    try:
        path_name =  Path(filename).suffix.lower()
        print(path_name)
        if "протокол" in filename.lower():
            print("ОБРАБОТКА: ", filename)
            print(text)
            extracted = extract_tender_fields(text, ["Победитель","Другие участники","Дата исполнения договора", "Филиал/РЭС"])
        else:
            extracted = {
                "Победитель": None,
                "Другие участники": None,
                "Ячейки": None,
                "Кол-во ячеек": None,
                "Типовой проект": None,
                "Проектировщик": None,
                "Дата исполнения договора": None,
                "Филиал/РЭС": None,
            }
        merge_extracted_into_accumulator(accumulator, extracted)
        return True
    except Exception as e:
        logger.exception(
            "EXTRACT_FIELDS ERROR | %s",
            format_log_kv(task_id=task_id, filename=filename, error=str(e))
        )
        return False


def read_supported_file_and_merge(target: Path, accumulator: dict, task_id: str | None = None) -> Dict:
    ext = target.suffix.lower()
    size_bytes = target.stat().st_size if target.exists() and target.is_file() else None
    started = time.perf_counter()

    logger.info(
        "READ_FILE START | %s",
        format_log_kv(task_id=task_id, path=target, filename=target.name, ext=ext, size_bytes=size_bytes)
    )

    try:
        text = read_file(target, task_id=task_id)
        elapsed = time.perf_counter() - started

        if not text.strip():
            logger.warning(
                "READ_FILE EMPTY | %s",
                format_log_kv(
                    task_id=task_id,
                    path=target,
                    filename=target.name,
                    ext=ext,
                    size_bytes=size_bytes,
                    elapsed=f"{elapsed:.3f}s",
                )
            )
            return {"ok": False, "error": "Текст не извлечен", "skipped": False}

        merged_ok = process_text_into_accumulator(
            text=text,
            accumulator=accumulator,
            filename=target.name,
            task_id=task_id,
        )

        del text

        logger.info(
            "READ_FILE DONE | %s",
            format_log_kv(
                task_id=task_id,
                path=target,
                filename=target.name,
                ext=ext,
                size_bytes=size_bytes,
                elapsed=f"{elapsed:.3f}s",
                merged=merged_ok,
            )
        )
        return {"ok": True, "error": None, "skipped": False}

    except ValueError as e:
        elapsed = time.perf_counter() - started
        logger.warning(
            "READ_FILE NO_TEXT | %s",
            format_log_kv(
                task_id=task_id,
                path=target,
                filename=target.name,
                ext=ext,
                size_bytes=size_bytes,
                elapsed=f"{elapsed:.3f}s",
                error=str(e),
            )
        )
        return {"ok": False, "error": str(e), "skipped": False}

    except Exception as e:
        elapsed = time.perf_counter() - started
        error_text = str(e)

        logger.exception(
            "READ_FILE ERROR | %s",
            format_log_kv(
                task_id=task_id,
                path=target,
                filename=target.name,
                ext=ext,
                size_bytes=size_bytes,
                elapsed=f"{elapsed:.3f}s",
                error=error_text,
            )
        )

        is_lo_timeout = "LibreOffice timeout" in error_text

        return {
            "ok": False,
            "error": error_text,
            "skipped": is_lo_timeout,
        }

def read_path_and_merge(path: str | Path, accumulator: dict, task_id: str | None = None) -> Dict:
    path = Path(path)

    logger.info("READ_PATH START | %s", format_log_kv(task_id=task_id, path=path, is_dir=path.is_dir() if path.exists() else None))

    stats = {
        "ok": 0,
        "failed": 0,
        "skipped": 0,
        "total": 0,
    }

    def walk(target: Path):
        if not target.exists():
            stats["failed"] += 1
            stats["total"] += 1
            return

        if target.is_dir():
            files = list(iter_candidate_files(target))
            logger.info("READ_DIR START | %s", format_log_kv(task_id=task_id, path=target, candidate_files=len(files)))

            if not files:
                stats["failed"] += 1
                stats["total"] += 1
                return

            regular_files = []
            archive_files = []

            for child in files:
                ext = child.suffix.lower()
                if ext in ARCHIVE_EXTENSIONS and not is_office_zip(child):
                    archive_files.append(child)
                else:
                    regular_files.append(child)

            logger.info(
                "READ_DIR SPLIT | %s",
                format_log_kv(
                    task_id=task_id,
                    path=target,
                    regular_files=len(regular_files),
                    archive_files=len(archive_files),
                    workers=1,
                )
            )

            for file_path in regular_files:
                stats["total"] += 1
                result = read_supported_file_and_merge(file_path, accumulator, task_id)
                if result["ok"]:
                    stats["ok"] += 1
                elif result.get("skipped"):
                    stats["skipped"] += 1
                else:
                    stats["failed"] += 1

            for archive_path in archive_files:
                logger.info("READ_DIR ARCHIVE_CHILD | %s", format_log_kv(task_id=task_id, parent=target, archive=archive_path))
                walk(archive_path)
            return

        ext = target.suffix.lower()

        try:
            if ext in ARCHIVE_EXTENSIONS and not is_office_zip(target):
                with tempfile.TemporaryDirectory() as tmp:
                    extracted_dir = Path(tmp)

                    if ext == ".zip":
                        extract_zip(target, extracted_dir, task_id=task_id)
                    elif ext == ".rar":
                        extract_rar(target, extracted_dir, task_id=task_id)

                    walk(extracted_dir)
                return

            if ext not in SUPPORTED_EXTENSIONS:
                stats["skipped"] += 1
                stats["total"] += 1
                return

            stats["total"] += 1
            result = read_supported_file_and_merge(target, accumulator, task_id)
            if result["ok"]:
                stats["ok"] += 1
            elif result.get("skipped"):
                stats["skipped"] += 1
            else:
                stats["failed"] += 1

        except Exception:
            stats["failed"] += 1
            stats["total"] += 1
            logger.exception("READ_PATH ITEM ERROR | %s", format_log_kv(task_id=task_id, path=target))

    walk(path)

    logger.info(
        "READ_PATH DONE | %s",
        format_log_kv(
            task_id=task_id,
            path=path,
            total=stats["total"],
            ok=stats["ok"],
            failed=stats["failed"],
            skipped=stats["skipped"],
        )
    )
    return stats


# =========================
# Multipart RAR
# =========================

def group_attached_files(attached_files: list) -> list:
    single = []

    part_scheme_groups = {}
    oldstyle_main = {}
    oldstyle_subparts = {}

    for doc in attached_files:
        filename = (doc.get("filename") or "").strip()
        info = get_rar_group_info(filename)

        if not info:
            single.append({
                "group_type": "single",
                "doc": doc,
            })
            continue

        if info["scheme"] == "part":
            part_scheme_groups.setdefault(info["base"], []).append({
                "doc": doc,
                "part_num": info["part_num"],
                "is_start": info["is_start"],
                "scheme": "part",
                "base": info["base"],
            })

        elif info["scheme"] == "oldstyle_main":
            oldstyle_main[info["base"]] = {
                "doc": doc,
                "part_num": info["part_num"],
                "is_start": info["is_start"],
                "scheme": "oldstyle",
                "base": info["base"],
            }

        elif info["scheme"] == "oldstyle_subpart":
            oldstyle_subparts.setdefault(info["base"], []).append({
                "doc": doc,
                "part_num": info["part_num"],
                "is_start": info["is_start"],
                "scheme": "oldstyle",
                "base": info["base"],
            })

    result = single[:]

    for base, parts in part_scheme_groups.items():
        parts.sort(key=lambda x: x["part_num"])
        result.append({
            "group_type": "multipart_rar",
            "base": base,
            "scheme": "part",
            "parts": parts,
        })

    processed_oldstyle_bases = set()

    for base, subparts in oldstyle_subparts.items():
        parts = []
        if base in oldstyle_main:
            parts.append(oldstyle_main[base])
        parts.extend(subparts)
        parts.sort(key=lambda x: x["part_num"])

        result.append({
            "group_type": "multipart_rar",
            "base": base,
            "scheme": "oldstyle",
            "parts": parts,
        })
        processed_oldstyle_bases.add(base)

    for base, main_doc in oldstyle_main.items():
        if base not in processed_oldstyle_bases:
            result.append({
                "group_type": "single",
                "doc": main_doc["doc"],
            })

    return result


def download_multipart_rar(parts: list, work_dir: Path, task_id: str | None = None) -> Path:
    downloaded_parts = []
    session = get_thread_session()

    logger.info(
        "MULTIPART_RAR START | %s",
        format_log_kv(task_id=task_id, work_dir=work_dir, parts=len(parts))
    )

    for item in parts:
        doc = item["doc"]
        logger.info(
            "MULTIPART_RAR PART START | %s",
            format_log_kv(
                task_id=task_id,
                filename=doc.get("filename"),
                url=doc["url"],
                part_num=item["part_num"],
                is_start=item["is_start"],
            )
        )

        downloaded = download_file(
            session=session,
            url=doc["url"],
            download_dir=work_dir,
            filename=doc.get("filename"),
            task_id=task_id,
        )

        logger.info(
            "MULTIPART_RAR PART DONE | %s",
            format_log_kv(
                task_id=task_id,
                filename=downloaded.name,
                part_num=item["part_num"],
                is_start=item["is_start"],
            )
        )

        downloaded_parts.append({
            "path": downloaded,
            "part_num": item["part_num"],
            "is_start": item["is_start"],
        })

    downloaded_parts.sort(key=lambda x: x["part_num"])

    start_parts = [x for x in downloaded_parts if x["is_start"]]
    chosen = start_parts[0]["path"] if start_parts else downloaded_parts[0]["path"]

    logger.info("MULTIPART_RAR DONE | %s", format_log_kv(task_id=task_id, chosen=chosen, total_parts=len(downloaded_parts)))
    return chosen


# =========================
# Интеграция
# =========================

def process_one_attached_file_and_merge(item: dict, tmp_dir: Path, accumulator: dict) -> dict:
    work_dir = None
    session = get_thread_session()
    started = time.perf_counter()

    try:
        group_type = item["group_type"]

        if group_type == "single":
            doc = item["doc"]
            source_url = doc["url"]
            source_name = doc.get("filename") or "file"
            task_id = short_id(source_url)

            work_dir = ensure_dir(
                tmp_dir / f"{safe_filename(source_name)}_{stable_id(source_url)}"
            )

            logger.info(
                "ATTACHED START | %s",
                format_log_kv(
                    task_id=task_id,
                    group_type=group_type,
                    source_name=source_name,
                    source_url=source_url,
                    work_dir=work_dir,
                )
            )

            downloaded_path = download_file(
                session=session,
                url=source_url,
                download_dir=work_dir,
                filename=doc.get("filename"),
                task_id=task_id,
            )

            stats = read_path_and_merge(downloaded_path, accumulator, task_id=task_id)

            elapsed = time.perf_counter() - started
            logger.info(
                "ATTACHED DONE | %s",
                format_log_kv(
                    task_id=task_id,
                    group_type=group_type,
                    source_url=source_url,
                    ok=stats["ok"],
                    failed=stats["failed"],
                    skipped=stats["skipped"],
                    elapsed=f"{elapsed:.3f}s",
                )
            )

            return {
                "source_url": source_url,
                "error": None,
                "stats": stats,
            }

        elif group_type == "multipart_rar":
            base = item["base"]
            parts = item["parts"]
            pseudo_url = " | ".join(p["doc"]["url"] for p in parts)
            task_id = short_id(pseudo_url)

            work_dir = ensure_dir(
                tmp_dir / f"{safe_filename(base)}_{stable_id(pseudo_url)}"
            )

            logger.info(
                "ATTACHED START | %s",
                format_log_kv(
                    task_id=task_id,
                    group_type=group_type,
                    base=base,
                    parts=len(parts),
                    work_dir=work_dir,
                )
            )

            first_part_path = download_multipart_rar(parts, work_dir, task_id=task_id)
            stats = read_path_and_merge(first_part_path, accumulator, task_id=task_id)

            elapsed = time.perf_counter() - started
            logger.info(
                "ATTACHED DONE | %s",
                format_log_kv(
                    task_id=task_id,
                    group_type=group_type,
                    base=base,
                    ok=stats["ok"],
                    failed=stats["failed"],
                    skipped=stats["skipped"],
                    elapsed=f"{elapsed:.3f}s",
                )
            )

            return {
                "source_url": pseudo_url,
                "error": None,
                "stats": stats,
            }

        else:
            raise RuntimeError(f"Неизвестный group_type: {group_type}")

    except requests.RequestException as e:
        logger.exception("ATTACHED DOWNLOAD ERROR")
        return {"source_url": item, "error": f"Ошибка скачивания: {e}", "stats": None}

    except Exception as e:
        logger.exception("ATTACHED PROCESS ERROR")
        return {"source_url": item, "error": f"Ошибка обработки: {e}", "stats": None}

    finally:
        try:
            if work_dir and work_dir.exists():
                logger.debug("WORKDIR CLEANUP | %s", format_log_kv(path=work_dir))
                shutil.rmtree(work_dir, ignore_errors=True)
        except Exception as cleanup_error:
            logger.warning("Не удалось удалить временную папку %s: %s", work_dir, cleanup_error)


def process_attached_files_and_merge(attached_files: list, tmp_dir: str | Path) -> dict:
    tmp_dir = ensure_dir(tmp_dir)
    accumulator = init_result_accumulator()

    if not attached_files:
        logger.info("ATTACHED BATCH SKIP | reason=no_attached_files")
        return finalize_result_accumulator(accumulator)

    grouped_items = group_attached_files(attached_files)

    logger.info(
        "ATTACHED BATCH START | %s",
        format_log_kv(
            groups=len(grouped_items),
            mode="streaming_merge",
            tmp_dir=tmp_dir,
        )
    )

    batch_started = time.perf_counter()

    for idx, item in enumerate(grouped_items, start=1):
        logger.info(
            "ATTACHED BATCH ITEM START | %s",
            format_log_kv(
                index=idx,
                total=len(grouped_items),
                group_type=item.get("group_type"),
                base=item.get("base"),
                filename=(item.get("doc") or {}).get("filename") if item.get("group_type") == "single" else None,
            )
        )

        item_started = time.perf_counter()
        process_one_attached_file_and_merge(item, tmp_dir, accumulator)
        item_elapsed = time.perf_counter() - item_started

        logger.info(
            "ATTACHED BATCH ITEM DONE | %s",
            format_log_kv(
                index=idx,
                total=len(grouped_items),
                group_type=item.get("group_type"),
                elapsed=f"{item_elapsed:.3f}s",
            )
        )

    elapsed = time.perf_counter() - batch_started
    logger.info(
        "ATTACHED BATCH DONE | %s",
        format_log_kv(
            groups=len(grouped_items),
            elapsed=f"{elapsed:.3f}s",
        )
    )

    return finalize_result_accumulator(accumulator)



if __name__ == "__main__":

    from pathlib import Path
    import json

    folder = Path("/app/app/goszakupki_requests/tmp/ул. Феодосийская, зу 7 (extract.me)")

    accumulator = init_result_accumulator()
    stats = read_path_and_merge(folder, accumulator, task_id="local_test")
    result = finalize_result_accumulator(accumulator)

    print("STATS:")
    print(json.dumps(stats, ensure_ascii=False, indent=2))

    print("\nRESULT:")
    print(json.dumps(result, ensure_ascii=False, indent=2))